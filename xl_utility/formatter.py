from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile

import re, string


def separate_names(col_names, excel_file):
    def _separated_name(col_name, list_of_test_payloads, ws, condition):
        _shared_not_text_exception(col_name, ws)
        
        if not _find_column_by_name("lastname", ws):
            ws.insert_cols(0)
            ws["A1"].value = "First Name"

            ws.insert_cols(0)
            ws["A1"].value = "Last Name"

        has_first = str(_find_column_by_name("firstname", ws))
        has_last = str(_find_column_by_name("lastname", ws))
        
        def _core(has_initial):
            split_cols = {
                "first": list(),
                "last": list()
            }
            for row_idx in condition:
                position = has_initial + str(int(row_idx)+1)
                cell = ws[position]
                split_name = str(cell.value).split()
                
                split_cols["first"].append(split_name[0])
                
                if len(split_name) > 1:
                    split_cols["last"].append(split_name[1])

            def _new_column(col_letter, col_position):
                altered_for_test = {}
                altered_for_test["column"] = ws[col_letter + "1"].value
                altered_for_test["data"] = list()

                for row_idx in condition:
                    cell = split_cols[col_position][int(row_idx)-1].title()
                    ws[col_letter + str(int(row_idx)+1)].value = cell
                    
                    altered_for_test["data"].append(cell)

                return altered_for_test

            list_of_test_payloads.append(_new_column(has_first, "first"))
            list_of_test_payloads.append(_new_column(has_last, "last"))
        
        _validate_column(_core, col_name, col_names, ws)
        ws.delete_cols(3)

    return _parse_sheet_data(col_names, _separated_name, excel_file)

def separate_addresses(col_names, excel_file):
    def _separated_address(col_name, list_of_test_payloads, ws, condition):
        _shared_has_number_exception(col_name, ws)
        
        def _alter_cell(row_idx, has_initial):
            cell = str(ws[has_initial + str(int(row_idx)+1)].value)
            temp = re.compile("([0-9]+)([a-zA-Z]+)")

            if len(cell.split()) > 1:
                return cell
            else:
                altered_cell = " ".join(map(lambda x: x.capitalize(), temp.match(cell).groups()))
                ws[has_initial + str(int(row_idx)+1)].value = altered_cell

                return altered_cell
                
        altered_for_test = _alter_sheet_data(_alter_cell, col_name, col_names, ws)
        list_of_test_payloads.append(altered_for_test)
        
    return _parse_sheet_data(col_names, _separated_address, excel_file)

def capitalize_firstLetter(col_names, excel_file):
    def _capitalized_first(col_name, list_of_test_payloads, ws, condition):
        _shared_has_number_exception(col_name, ws)

        def _alter_cell(row_idx, has_initial):
            cell = str(ws[has_initial + str(int(row_idx)+1)].value)
            altered_cell = cell.title()
            ws[has_initial + str(int(row_idx)+1)].value = altered_cell

            return altered_cell

        altered_for_test = _alter_sheet_data(_alter_cell, col_name, col_names, ws)
        list_of_test_payloads.append(altered_for_test)

    return _parse_sheet_data(col_names, _capitalized_first, excel_file)

def capitalize_all(col_names, excel_file):
    def _capitalize_all(col_name, list_of_test_payloads, ws, condition):
        _shared_has_number_exception(col_name, ws)
        
        def _alter_cell(row_idx, has_initial):
            cell = str(ws[has_initial + str(int(row_idx)+1)].value)
            altered_cell = cell.upper()
            ws[has_initial + str(int(row_idx)+1)].value = altered_cell

            return altered_cell

        altered_for_test = _alter_sheet_data(_alter_cell, col_name, col_names, ws)
        list_of_test_payloads.append(altered_for_test)
    
    return _parse_sheet_data(col_names, _capitalize_all, excel_file)



# Private functions

def _shared_not_text_exception(col_name, ws):
    temp = re.compile("([a-zA-Z]+)")
    if not temp.match(_clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)):
        raise TypeError("Non text cells are forbidden in this function.")

def _shared_has_text_exception(col_name, ws):
    temp = re.compile("([a-zA-Z]+)")
    if temp.match(_clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)):
        raise TypeError("Text cells are forbidden in this function.")

def _shared_has_number_exception(col_name, ws):
    cleaned_cell = _clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)
    if cleaned_cell.isdigit():
        raise TypeError("Number cells are forbidden in this function.")    

def _clean_String(string):
    new_string = ''.join(e for e in string if e.isalnum())

    return new_string

def _validate_column(_core, col_name, col_names, ws):
    has_initial = _find_column_by_name(col_name, ws)

    if has_initial and ws[has_initial + "1"].value.replace(" ", "").lower() in list(map(lambda x: re.sub(r"[^a-zA-Z0-9 ]", "", x.replace(" ", "").lower()), col_names)):   
        return _core(has_initial)
    else:
        raise KeyError("Column name not found.")

def _alter_sheet_data(_alter_cell, col_name, col_names, ws):
    def _core(has_column_letter):
        altered_for_test = {}
        altered_for_test["column"] = ws[has_column_letter + "1"].value
        altered_for_test["data"] = list()

        for row_idx in range(1, ws.max_row):
            altered_cell = _alter_cell(row_idx, has_column_letter)
            ws[has_column_letter + str(int(row_idx)+1)].value = altered_cell
            altered_for_test["data"].append(altered_cell)

        return altered_for_test

    return _validate_column(_core, col_name, col_names, ws)

def _find_column_by_name(name, ws):
    col_names = list(map(lambda x: x.replace(" ", "").lower(), list(map(lambda x: x.value, list(ws.iter_rows())[0]))))
    name_payload = re.sub(r"[^a-zA-Z0-9 ]", "", name.replace(" ", "").lower())
    
    return None if name_payload not in col_names else get_column_letter(col_names.index(name_payload)+1)

def _parse_sheet_data(col_names, handle_alterations, excel_file):
    if type(col_names) is not list or not all(list(map(lambda x: type(x) == str, col_names))):
        raise TypeError("Arg with type {} is not list of strings.".format(type(col_names)))

    payload = {
        "test_list": list(),
        "buffer": None,
        "exception": ""
    }

    wb = load_workbook(filename=excel_file, data_only=True)
    ws = wb.active
    
    condition = range(1, ws.max_row)
    
    # Filters column names
    for name in col_names:
        try:
            handle_alterations(name, payload["test_list"], ws, condition)
        except Exception as error:
            payload["exception"] = payload["exception"] + str(error)+" -{} is rejected! ".format(name)

            continue

    if not len(payload["test_list"]):
        return payload 
    
    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()
    payload["buffer"] = stream

    return payload

    