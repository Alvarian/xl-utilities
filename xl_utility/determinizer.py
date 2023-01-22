from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile, _TemporaryFileWrapper

from io import BytesIO
import re, string, csv, io



# def guess_gender(col_names, excel_file):
#     return "guess"

# def generate_uuid(col_names, excel_file):
#     return "genID"

def insert_mock_data(excel_file, col_names, mock_data):
    # Given column and given data(array of, or single), if column not exist create column, if array - iter through into each row, if single - clone through into each row
    def _insert(col_name, list_of_test_payloads, ws, condition):
        if not _find_column_by_name(col_name, ws):
            ws.insert_cols(0)
            ws["A1"].value = col_name

        has_column_letter = str(_find_column_by_name(col_name, ws))

        if type(mock_data) == list and len(mock_data) < condition:
            raise AssertionError("Mock data list given is less than total rows. ")

        altered_for_test = {
            "column": ws[has_column_letter + "1"].value,
            "data": list()
        }
        
        mock_iter = len(mock_data) if type(mock_data) == list else condition
        for idx in range(1, mock_iter):
            new_cell = mock_data[idx] if type(mock_data) == list else mock_data
            ws[has_column_letter + str(int(idx)+1)].value = new_cell
            altered_for_test["data"].append(new_cell)
            
        list_of_test_payloads.append(altered_for_test)
        
    return _parse_sheet_data(col_names, _insert, excel_file)

# def validate_emails(col_names, excel_file):
#     # Creates new column next to email column specifying per row if email is valid. If invalid, if dup - have cell instruct the other duplicates, if not email - have cell indicate value error
#     return "valEmails"


# Exception guards
def _shared_not_text_exception(col_name, ws):
    temp = re.compile("([a-zA-Z]+)")
    if not temp.match(_clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)):
        raise TypeError("Non text cells are forbidden in this function.")

def _shared_has_text_exception(col_name, ws):
    temp = re.compile("([a-zA-Z]+)")
    if temp.match(_clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)):
        raise TypeError("Text cells are forbidden in this function.")

def _shared_has_number_exception(col_name, ws):
    cleaned_cell = _clean_String(ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value)
    if cleaned_cell and cleaned_cell.isdigit():
        raise TypeError("Number cells are forbidden in this function.")    

def _shared_has_email_exception(col_name, ws):
    pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    if re.match(pat, ws[_find_column_by_name(col_name.replace(" ", "").lower(), ws)+"2"].value):
        raise TypeError("Email cells are forbidden in this function.")    


# Private functions
def _clean_String(string):
    new_string = ''.join(e for e in string if e.isalnum()) if string else None

    return new_string

def _validate_column(_core, col_name, col_names, ws):
    has_initial = _find_column_by_name(col_name, ws)

    if has_initial and ws[has_initial + "1"].value.replace(" ", "").lower() in list(map(lambda x: re.sub(r"[^a-zA-Z0-9 ]", "", x.replace(" ", "").lower()), col_names)):   
        return _core(has_initial)
    else:
        raise KeyError("Column name not found.")

def _alter_sheet_data(_alter_cell, col_name, col_names, ws):
    def _core(has_column_letter):
        altered_for_test = {}
        altered_for_test["column"] = ws[has_column_letter + "1"].value
        altered_for_test["data"] = list()

        for row_idx in range(1, ws.max_row):
            altered_cell = _alter_cell(row_idx, has_column_letter)
            ws[has_column_letter + str(int(row_idx)+1)].value = altered_cell
            altered_for_test["data"].append(altered_cell)

        return altered_for_test

    return _validate_column(_core, col_name, col_names, ws)

def _find_column_by_name(name, ws):
    col_names = list(map(lambda x: x.replace(" ", "").lower(), list(map(lambda x: x.value, list(ws.iter_rows())[0]))))
    name_payload = re.sub(r"[^a-zA-Z0-9 ]", "", name.replace(" ", "").lower())
    
    return None if name_payload not in col_names else get_column_letter(col_names.index(name_payload)+1)

def _parse_sheet_data(col_names, handle_alterations, excel_file):
    if type(col_names) is not list or not all(list(map(lambda x: type(x) == str, col_names))):
        raise TypeError("Arg with type {} is not list of strings.".format(type(col_names)))
    
    payload = {
        "test_list": list(),
        "buffer": None,
        "exception": ""
    }
    
    inbound_buffer = excel_file["buffer"]
    file_name = inbound_buffer.name if isinstance(inbound_buffer, io.BufferedReader) or isinstance(inbound_buffer, _TemporaryFileWrapper) else inbound_buffer.filename
    file_data = inbound_buffer.read() if isinstance(inbound_buffer, io.BufferedReader) or isinstance(inbound_buffer, _TemporaryFileWrapper) else inbound_buffer.stream.read()
    if file_name.split('.')[-1] == "csv":
        stream = io.StringIO(file_data.decode("utf-8-sig"), newline=None)
        rows = list(filter(lambda x: x[0] != "", list(csv.reader(stream))))
        condition = len(rows)
        
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "A Snazzy Title"
        
        for row_idx in range(0, len(rows)):
            row = rows[row_idx]

            for column_index in range(0, len(row)):
                cell = row[column_index]
                column_letter = get_column_letter((column_index + 1))
                ws[column_letter + str(int(row_idx)+1)].value = cell

        # Filters column names
        for name in col_names:
            try:
                handle_alterations(name, payload["test_list"], ws, condition)
            except Exception as error:
                payload["exception"] = payload["exception"]+str(error)+" -{} is rejected! ".format(name)

                continue
            
        if not len(payload["test_list"]):
            return payload 
            
        tmp = NamedTemporaryFile()
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp
        payload["buffer"] = stream

        return payload
    else:
        wb = load_workbook(filename=BytesIO(file_data), data_only=True)
        ws = wb.active
        
        condition = ws.max_row
        
        # Filters column names
        for name in col_names:
            try:
                handle_alterations(name, payload["test_list"], ws, condition)
            except Exception as error:
                payload["exception"] = payload["exception"] + str(error)+" -{} is rejected! ".format(name)

                continue

        if not len(payload["test_list"]):
            return payload 
        
        tmp = NamedTemporaryFile()
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp
        payload["buffer"] = stream
        
        return payload
    